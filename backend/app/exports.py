"""Excel output. Column widths and a frozen header row, because these sheets
get opened on a phone as often as a laptop.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="14202C")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
WIDTHS = {"#": 5, "Job title": 46, "Client": 26, "Platform": 14, "Job link": 52,
          "Already tried by": 22, "Match ref": 30, "Metric": 34, "Value": 14,
          "BD": 20, "Applied by": 30, "People": 9, "Status": 12,
          "Found by": 24, "Description link": 46,
          "When · ET": 18, "Applied as": 20, "Developer": 20, "Role": 40,
          "Round": 16, "Outcome": 12, "Follows": 18, "Reported by": 20,
          "What": 40, "Due · ET": 18, "Set by": 20, "Submitted": 18,
          "Submission": 46, "Notes": 60, "Applied on": 18}


# Excel treats a leading =, +, - or @ as the start of a formula. These sheets
# are built from cells somebody typed into their own spreadsheet, then opened by
# a colleague — so a title of `=cmd|'/c calc'!A1` would execute on their machine,
# not ours. A leading apostrophe tells Excel "this is text", and is not shown.
_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(value):
    if isinstance(value, str) and value.startswith(_FORMULA_LEAD):
        return "'" + value
    return value


def _write_sheet(book: Workbook, title: str, columns: list[str], rows: list[list]) -> None:
    sheet = book.create_sheet(title[:31] or "Sheet")
    sheet.append(columns)
    for cell in sheet[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    for row in rows:
        sheet.append([_safe_cell(value) for value in row])
    for index, column in enumerate(columns, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = WIDTHS.get(column, 18)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def assignment_workbook(person: str, jobs: list[dict]) -> bytes:
    """One profile's list, as the file people actually work from.

    Carries `Found by` and the description link, because the download is what
    somebody works through on a train and it should not be a worse document
    than the screen it came from. Knowing a run of jobs all came off one
    colleague's search is what tells you whether to trust the next twenty.
    """
    book = Workbook()
    book.remove(book.active)
    columns = ["#", "Job title", "Client", "Found by", "Platform", "Job link",
               "Description link", "Status"]
    rows = [
        [i, j.get("title", ""), j.get("company", ""),
         ", ".join(j.get("found_by") or []), j.get("platform", ""),
         j.get("url", ""), j.get("description_url", ""), j.get("status", "pending")]
        for i, j in enumerate(jobs, start=1)
    ]
    _write_sheet(book, f"{person} jobs", columns, rows)
    stream = io.BytesIO()
    book.save(stream)
    return stream.getvalue()


def pipeline_workbook(interviews: list[dict], assessments: list[dict]) -> bytes:
    """Every conversation and every take-home, in one file.

    The half of this product that has no export at all until now. Everything
    downloadable so far counts what went out — rows, lists, duplication — and a
    team can improve all of it without winning a single piece of work. This is
    the other half, and it is the part somebody wants in a spreadsheet when
    they are asked what the last quarter actually produced.

    Times are the team's clock, said so in the column heading, because a
    spreadsheet has no tooltip to explain itself later.
    """
    book = Workbook()
    book.remove(book.active)

    _write_sheet(
        book, "Conversations",
        ["When · ET", "Applied as", "Developer", "Client", "Role", "Round",
         "Follows", "How", "Status", "Outcome", "Reported by", "Job link"],
        [[row["when"]["label"], row["profile"], row.get("developer") or "",
          row["client"], row["role"],
          f"{row['round']} of {row['rounds']} · {row['stage']}"
          if row.get("rounds", 1) > 1 else row["stage"],
          (row["follows"]["stage"] if row.get("follows") else ""),
          row["mode"], row["status"], row["outcome"],
          row.get("reported_by") or "",
          (row["job"] or {}).get("url", "")]
         for row in interviews])

    _write_sheet(
        book, "Take-homes",
        ["What", "Client", "Applied as", "Developer", "Due · ET", "Status",
         "Set by", "Submitted", "Submission", "Job link", "Applied on", "Notes"],
        [[row["title"], row["client"], row["profile"], row.get("developer") or "",
          (row["due"]["label"] if row.get("due") else "none set"),
          # Said in words rather than left as a status somebody has to know how
          # to read. A file outlives the screen that explains it.
          f"{row['status']} · overdue" if row["overdue"] else row["status"],
          row.get("set_by") or "",
          (row["submitted"]["label"] if row.get("submitted") else ""),
          row["submission_url"],
          # The posting, on the same terms as the Conversations sheet. A
          # spreadsheet that describes one conversation two different ways is
          # the thing this whole payload is trying not to be.
          (row["job"] or {}).get("url", ""),
          (row["job"] or {}).get("applied_on", ""),
          row["notes"]]
         for row in assessments])

    stream = io.BytesIO()
    book.save(stream)
    return stream.getvalue()


def report_workbook(report: dict, per_person: dict[str, list[dict]],
                    collisions: list[dict], matrix_names: list[str],
                    matrix: list[list[int]]) -> bytes:
    book = Workbook()
    book.remove(book.active)

    _write_sheet(book, "Summary", ["Metric", "Value"],
                 [[k, v] for k, v in report.items()])

    _write_sheet(book, "Collisions",
                 ["Job title", "Client", "Platform", "Job link", "Applied by", "People"],
                 [[c["title"], c["company"], c["platform"], c["url"],
                   ", ".join(c["applied_by"]), len(c["applied_by"])] for c in collisions])

    _write_sheet(book, "Overlap matrix", ["BD"] + matrix_names,
                 [[matrix_names[i]] + row for i, row in enumerate(matrix)])

    for person, jobs in per_person.items():
        _write_sheet(book, person, ["#", "Job title", "Client", "Platform", "Job link"],
                     [[i, j.get("title", ""), j.get("company", ""),
                       j.get("platform", ""), j.get("url", "")]
                      for i, j in enumerate(jobs, start=1)])

    stream = io.BytesIO()
    book.save(stream)
    return stream.getvalue()
