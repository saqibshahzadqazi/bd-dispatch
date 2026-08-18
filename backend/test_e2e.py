"""Full round trip through the API, no running server needed.

    python seed.py --samples
    pytest test_e2e.py -v -s

The scenario throughout is the one the product exists for: two profiles with
the same skills, run by two different people. Khuram logged 30 jobs, Zahid 50,
ten of them the same posting found twice. Khuram gets back the 40 it has never
seen, Zahid the 20 it has never seen — and either may receive a job the other
has already used, because they are two candidates, not one person applying
twice.

Note on isolation: every test mints its own profiles. Unworked jobs are carried
between cycles on purpose, so tests that shared profiles would inherit each
other's open lists.
"""
import io
import itertools
import os
from pathlib import Path

import pytest
from fastapi.testclient import TestClient

SHEETS = Path(__file__).parent / "sample_sheets"

PEOPLE = [("ali@example.com", "Ali Raza"),
          ("sara@example.com", "Sara Khan"),
          ("hina@example.com", "Hina Malik")]

SEEDED = [("Khuram", "AI Engineer", "ali@example.com", "khuram-applied.csv"),
          ("Zahid", "AI Engineer", "sara@example.com", "zahid-applied.csv"),
          ("Nadia", "Full Stack Engineer", "hina@example.com", "nadia-applied.csv")]

_serial = itertools.count(1)


@pytest.fixture(scope="module")
def client(tmp_path_factory):
    os.environ["DATABASE_URL"] = f"sqlite:///{tmp_path_factory.mktemp('db')}/test.db"
    from app.main import SessionLocal, app, engine, hash_password  # noqa: E402
    from app.models import Base, Profile, User  # noqa: E402

    Base.metadata.create_all(engine)
    db = SessionLocal()
    db.add(User(email="admin@example.com", name="Manager",
                password_hash=hash_password("admin12345"), role="admin"))
    for email, name in PEOPLE:
        db.add(User(email=email, name=name,
                    password_hash=hash_password("bdpass12345"), role="bd"))
    db.commit()
    owners = {u.email: u.id for u in db.query(User).all()}
    for name, headline, email, _ in SEEDED:
        db.add(Profile(name=name, headline=headline, platform="Upwork",
                       user_id=owners[email]))
    db.commit()
    db.close()
    return TestClient(app)


def token(client, email, password="bdpass12345"):
    response = client.post("/api/auth/login", json={"email": email, "password": password})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['token']}"}


@pytest.fixture(scope="module")
def admin(client):
    return token(client, "admin@example.com", "admin12345")


@pytest.fixture(scope="module")
def seeded(client, admin):
    return {p["name"]: p["id"] for p in client.get("/api/profiles", headers=admin).json()}


@pytest.fixture
def new_profile(client, admin):
    """Mint a profile nothing else in the suite has touched."""
    users = {u["email"]: u["id"] for u in client.get("/api/users", headers=admin).json()}

    def make(email, headline="AI Engineer"):
        name = f"Ident{next(_serial)}"
        response = client.post("/api/profiles",
                               json={"name": name, "headline": headline,
                                     "platform": "Upwork", "user_id": users[email]},
                               headers=admin)
        assert response.status_code == 201, response.text
        return response.json()["id"]

    return make


def sheet(job_ids, tracking=False):
    """A CSV of Upwork postings, optionally with a referral tail on every link."""
    out = io.StringIO()
    out.write("Job Title,Company,Job URL\n")
    for n in job_ids:
        url = f"https://www.upwork.com/jobs/~01{n:016x}"
        if tracking:
            url += "?utm_source=email&referrer=digest"
        out.write(f"Role {n},Client {n % 12} Ltd,{url}\n")
    return out.getvalue().encode()


def open_cycle(client, admin, name, mode="cover", quota=500, one_per_client=False,
               auto_build_minutes=0):
    """Cycles default to no timer here so tests build when they mean to."""
    response = client.post("/api/batches", json={"name": name, "mode": mode, "quota": quota,
                                                 "one_per_client": one_per_client,
                                                 "auto_build_minutes": auto_build_minutes},
                           headers=admin)
    assert response.status_code == 201, response.text
    return response.json()["id"]


def hand_in(client, batch_id, headers, profile_id, payload, filename="s.csv"):
    response = client.post(f"/api/batches/{batch_id}/uploads",
                           data={"profile_id": profile_id},
                           files={"file": (filename, payload, "text/csv")}, headers=headers)
    assert response.status_code == 201, response.text
    return response.json()


def listing(client, batch_id, profile_id, headers):
    return client.get(f"/api/batches/{batch_id}/profiles/{profile_id}/sheet",
                      headers=headers).json()["jobs"]


# --------------------------------------------------------------------------- #

def test_the_headline_scenario(client, admin, seeded):
    """Khuram logged 30, Zahid logged 50, ten shared. 40 back and 20 back."""
    if not (SHEETS / "khuram-applied.csv").exists():
        pytest.skip("Run `python seed.py --samples` first.")

    batch_id = open_cycle(client, admin, "Week 32")
    for name, _, email, filename in SEEDED[:2]:
        body = hand_in(client, batch_id, token(client, email), seeded[name],
                       (SHEETS / filename).read_bytes(), filename)
        assert body["mapping"]["url"], f"auto-mapper missed the link column in {filename}"
        print(f"  {name:8s} {body['row_count']:3d} rows  mapped -> {body['mapping']}")

    data = client.post(f"/api/batches/{batch_id}/compute", headers=admin).json()

    print("\n  Report:")
    for key, value in data["report"].items():
        print(f"    {key:38s} {value}")
    print("\n  New lists:")
    for person in data["participants"]:
        print(f"    {person['name']:8s} {person['headline']:22s} "
              f"{person['person']:12s} {person['assigned']:4d} jobs")

    handed = {p["name"]: p["assigned"] for p in data["participants"]}
    assert handed == {"Khuram": 40, "Zahid": 20}
    assert data["report"]["Jobs two profiles both applied to"] == 10

    xlsx = client.get(f"/api/batches/{batch_id}/profiles/{seeded['Khuram']}/sheet.xlsx",
                      headers=token(client, "ali@example.com"))
    assert xlsx.status_code == 200 and xlsx.content[:2] == b"PK"
    workbook = client.get(f"/api/batches/{batch_id}/report.xlsx", headers=admin)
    assert workbook.status_code == 200 and workbook.content[:2] == b"PK"
    print("  Excel downloads OK.")


def test_a_messy_sheet_still_maps(client, admin, seeded):
    """Nadia's sheet has different column names, blank links and shouty titles."""
    if not (SHEETS / "nadia-applied.csv").exists():
        pytest.skip("Run `python seed.py --samples` first.")
    batch_id = open_cycle(client, admin, "Messy")
    body = hand_in(client, batch_id, token(client, "hina@example.com"), seeded["Nadia"],
                   (SHEETS / "nadia-applied.csv").read_bytes(), "nadia-applied.csv")
    assert body["mapping"] == {"url": "Post Link", "title": "Position Applied",
                               "company": "Client Name", "platform": "Job Portal",
                               "date": "Apply Date"}


def test_two_profiles_may_receive_the_same_job(client, admin, new_profile):
    """The behaviour the whole redesign is for: one posting, two candidates."""
    khuram = new_profile("ali@example.com")
    zahid = new_profile("sara@example.com")
    scout = new_profile("hina@example.com", "Full Stack Engineer")

    batch_id = open_cycle(client, admin, "Shared")
    hand_in(client, batch_id, token(client, "hina@example.com"), scout, sheet(range(2000, 2010)))
    hand_in(client, batch_id, token(client, "ali@example.com"), khuram, sheet(range(2100, 2105)))
    hand_in(client, batch_id, token(client, "sara@example.com"), zahid, sheet(range(2200, 2205)))
    assert client.post(f"/api/batches/{batch_id}/compute", headers=admin).status_code == 200

    a = {j["url"] for j in listing(client, batch_id, khuram, token(client, "ali@example.com"))}
    b = {j["url"] for j in listing(client, batch_id, zahid, token(client, "sara@example.com"))}
    assert len(a & b) == 10, "both should be offered the ten jobs neither has touched"


def test_rebuilding_gives_the_same_answer(client, admin, new_profile):
    khuram = new_profile("ali@example.com")
    zahid = new_profile("sara@example.com")
    batch_id = open_cycle(client, admin, "Rebuild")
    hand_in(client, batch_id, token(client, "ali@example.com"), khuram, sheet(range(1000, 1020)))
    hand_in(client, batch_id, token(client, "sara@example.com"), zahid, sheet(range(1010, 1040)))

    first = client.post(f"/api/batches/{batch_id}/compute", headers=admin).json()
    second = client.post(f"/api/batches/{batch_id}/compute", headers=admin).json()
    assert second["report"] == first["report"], "rebuilding changed the outcome"
    assert ([p["assigned"] for p in second["participants"]]
            == [p["assigned"] for p in first["participants"]])


def test_unworked_jobs_come_back_next_cycle(client, admin, new_profile):
    khuram = new_profile("ali@example.com")
    zahid = new_profile("sara@example.com")
    ali, sara = token(client, "ali@example.com"), token(client, "sara@example.com")

    first = open_cycle(client, admin, "Carry 1")
    hand_in(client, first, ali, khuram, sheet(range(3000, 3010)))
    hand_in(client, first, sara, zahid, sheet(range(3010, 3020)))
    client.post(f"/api/batches/{first}/compute", headers=admin)

    mine = listing(client, first, khuram, ali)
    assert len(mine) == 10, "the ten Zahid found and Khuram has not tried"
    client.patch(f"/api/assignments/{mine[0]['id']}", json={"status": "applied"}, headers=ali)
    client.patch(f"/api/assignments/{mine[1]['id']}", json={"status": "skipped"}, headers=ali)

    # Nothing new enters the pool second time round, so only carry-forward can
    # fill a list.
    second = open_cycle(client, admin, "Carry 2")
    hand_in(client, second, ali, khuram, sheet(range(3000, 3010)))
    hand_in(client, second, sara, zahid, sheet(range(3010, 3020)))
    data = client.post(f"/api/batches/{second}/compute", headers=admin).json()

    handed = {p["id"]: p["assigned"] for p in data["participants"]}
    assert handed[khuram] == 8, "the eight still open return; applied and skipped do not"

    stale = listing(client, first, khuram, ali)
    assert len(stale) == 2, "work that moved forward should not stay open on the old cycle"


def test_split_mode_still_gives_each_job_to_one_profile(client, admin, new_profile):
    khuram = new_profile("ali@example.com")
    zahid = new_profile("sara@example.com")
    batch_id = open_cycle(client, admin, "Split", mode="split")
    hand_in(client, batch_id, token(client, "ali@example.com"), khuram, sheet(range(4000, 4020)))
    hand_in(client, batch_id, token(client, "sara@example.com"), zahid, sheet(range(4020, 4040)))
    assert client.post(f"/api/batches/{batch_id}/compute", headers=admin).status_code == 200

    seen = set()
    for profile_id, email in ((khuram, "ali@example.com"), (zahid, "sara@example.com")):
        for job in listing(client, batch_id, profile_id, token(client, email)):
            assert job["url"] not in seen, f"{job['url']} went to two profiles in a split cycle"
            seen.add(job["url"])
    assert seen, "a split cycle should still dispatch something"


def test_one_person_can_run_two_profiles_independently(client, admin, new_profile):
    """What one identity has applied to must not restrict the other, even when
    the same person runs both."""
    first = new_profile("ali@example.com")
    second = new_profile("ali@example.com")
    ali = token(client, "ali@example.com")

    mine = {p["id"] for p in client.get("/api/profiles", headers=ali).json()}
    assert {first, second} <= mine

    batch_id = open_cycle(client, admin, "Two hats")
    hand_in(client, batch_id, ali, first, sheet(range(5000, 5010)))
    hand_in(client, batch_id, ali, second, sheet(range(5010, 5020)))
    data = client.post(f"/api/batches/{batch_id}/compute", headers=admin).json()

    handed = {p["id"]: p["assigned"] for p in data["participants"]}
    assert handed[first] == 10 and handed[second] == 10, \
        "each identity is judged on its own history, not its operator's"

    sheets = client.get(f"/api/batches/{batch_id}/my-sheets", headers=ali).json()
    assert {p["id"] for p in sheets["profiles"]} >= {first, second}


def test_jobs_can_be_typed_in_instead_of_uploaded(client, admin, new_profile):
    """A BD with no spreadsheet types rows on screen; they must reach the cycle
    exactly as an uploaded sheet would."""
    typed = new_profile("ali@example.com")
    other = new_profile("sara@example.com")
    ali = token(client, "ali@example.com")

    batch_id = open_cycle(client, admin, "Typed")
    empty = client.get(f"/api/batches/{batch_id}/profiles/{typed}/entries", headers=ali)
    assert empty.status_code == 200 and empty.json()["rows"] == []

    rows = [{"url": f"https://www.upwork.com/jobs/~01{n:016x}",
             "title": f"Role {n}", "company": f"Client {n}",
             "platform": "Upwork", "date": "2026-08-01"} for n in range(9000, 9006)]
    rows.append({"url": "", "title": "", "company": "", "platform": "", "date": ""})
    saved = client.put(f"/api/batches/{batch_id}/profiles/{typed}/entries",
                       json={"rows": rows}, headers=ali)
    assert saved.status_code == 200, saved.text
    assert saved.json()["row_count"] == 6, "the blank row should not be stored"

    # It reads back, and the manager sees it as a handed-in sheet.
    back = client.get(f"/api/batches/{batch_id}/profiles/{typed}/entries", headers=ali).json()
    assert len(back["rows"]) == 6 and back["typed"] is True
    handed = client.get(f"/api/batches/{batch_id}", headers=admin).json()["uploads"]
    assert any(u["profile_id"] == typed and u["row_count"] == 6 for u in handed)

    # And it dispatches like any other sheet.
    hand_in(client, batch_id, token(client, "sara@example.com"), other, sheet(range(9100, 9110)))
    data = client.post(f"/api/batches/{batch_id}/compute", headers=admin).json()
    handed = {p["id"]: p["assigned"] for p in data["participants"]}
    assert handed[typed] == 10, "the typed profile gets the ten it has not tried"
    assert handed[other] == 6, "the other profile gets the six that were typed in"


def test_opening_the_entry_screen_does_not_hand_anything_in(client, admin, new_profile):
    """Clicking into the table and away must not leave an empty sheet behind —
    the manager would read that as a profile having reported in."""
    typed = new_profile("ali@example.com")
    ali = token(client, "ali@example.com")
    batch_id = open_cycle(client, admin, "Typed empty")

    blank = client.put(f"/api/batches/{batch_id}/profiles/{typed}/entries",
                       json={"rows": [{"url": "", "title": "", "company": ""}]}, headers=ali)
    assert blank.status_code == 200 and blank.json()["row_count"] == 0
    assert client.get(f"/api/batches/{batch_id}", headers=admin).json()["uploads"] == []

    # Every new row arrives pre-stamped with the time, so a row carrying only a
    # timestamp is still an empty row.
    stamped = client.put(f"/api/batches/{batch_id}/profiles/{typed}/entries", headers=ali,
                         json={"rows": [{"url": "", "title": "", "company": "",
                                         "platform": "Upwork", "date": "2026-08-18 11:45"}]})
    assert stamped.status_code == 200 and stamped.json()["row_count"] == 0, \
        "a row with only a stamp and a platform is not a job"
    assert client.get(f"/api/batches/{batch_id}", headers=admin).json()["uploads"] == []

    # And clearing a real sheet takes it back out again.
    client.put(f"/api/batches/{batch_id}/profiles/{typed}/entries",
               json={"rows": [{"url": "https://www.upwork.com/jobs/~01000000000000beef"}]},
               headers=ali)
    assert len(client.get(f"/api/batches/{batch_id}", headers=admin).json()["uploads"]) == 1
    client.put(f"/api/batches/{batch_id}/profiles/{typed}/entries", json={"rows": []}, headers=ali)
    assert client.get(f"/api/batches/{batch_id}", headers=admin).json()["uploads"] == []


def test_the_applied_date_is_kept(client, admin, new_profile):
    """The date column used to be read off the sheet and thrown away. It should
    reach the history row, and marking a job applied should stamp it."""
    from sqlalchemy import select as sa_select

    from app.main import SessionLocal
    from app.models import Application

    one = new_profile("ali@example.com")
    two = new_profile("sara@example.com")
    ali = token(client, "ali@example.com")

    batch_id = open_cycle(client, admin, "Dates")
    client.put(f"/api/batches/{batch_id}/profiles/{one}/entries", headers=ali, json={"rows": [
        {"url": "https://www.upwork.com/jobs/~01000000000000d1d1", "title": "Dated",
         "company": "Acme", "platform": "Upwork", "date": "2026-08-17 09:15"}]})
    hand_in(client, batch_id, token(client, "sara@example.com"), two, sheet(range(9400, 9405)))
    client.post(f"/api/batches/{batch_id}/compute", headers=admin)

    db = SessionLocal()
    try:
        typed = db.scalar(sa_select(Application).where(Application.profile_id == one))
        assert typed.applied_on == "2026-08-17 09:15", "the typed date should be stored verbatim"

        # Marking something applied stamps the moment it happened.
        job = client.get(f"/api/batches/{batch_id}/profiles/{one}/sheet", headers=ali).json()["jobs"][0]
        client.patch(f"/api/assignments/{job['id']}", json={"status": "applied"}, headers=ali)
        db.expire_all()
        marked = db.scalars(sa_select(Application).where(Application.profile_id == one)).all()
        stamps = [a.applied_on for a in marked if a.applied_on]
        assert len(stamps) == 2 and any(s != "2026-08-17 09:15" for s in stamps)
    finally:
        db.close()


def test_typed_entries_reject_a_dangerous_link(client, admin, new_profile):
    typed = new_profile("ali@example.com")
    ali = token(client, "ali@example.com")
    batch_id = open_cycle(client, admin, "Typed hostile")
    client.put(f"/api/batches/{batch_id}/profiles/{typed}/entries",
               json={"rows": [{"url": "javascript:alert(1)", "title": "X", "company": "Y"}]},
               headers=ali)
    back = client.get(f"/api/batches/{batch_id}/profiles/{typed}/entries", headers=ali).json()
    assert back["rows"][0]["url"] == "", "a javascript: link must not survive being typed in"


def test_a_bd_cannot_type_into_someone_elses_profile(client, admin, seeded):
    batch_id = open_cycle(client, admin, "Typed boundary")
    denied = client.put(f"/api/batches/{batch_id}/profiles/{seeded['Zahid']}/entries",
                        json={"rows": [{"url": "https://x.com/jobs/1234567"}]},
                        headers=token(client, "ali@example.com"))
    assert denied.status_code == 403


def test_building_keeps_the_cycle_open_but_closing_shuts_it(client, admin, new_profile):
    """Lists are rebuilt on a timer while a cycle runs, so a build must not lock
    anyone out. Only closing the cycle does that."""
    one = new_profile("ali@example.com")
    two = new_profile("sara@example.com")
    ali = token(client, "ali@example.com")
    batch_id = open_cycle(client, admin, "Still open")
    hand_in(client, batch_id, ali, one, sheet(range(9200, 9205)))
    hand_in(client, batch_id, token(client, "sara@example.com"), two, sheet(range(9205, 9210)))
    built = client.post(f"/api/batches/{batch_id}/compute", headers=admin).json()
    assert built["status"] == "open" and built["last_built_at"]

    more = client.put(f"/api/batches/{batch_id}/profiles/{one}/entries",
                      json={"rows": [{"url": "https://www.upwork.com/jobs/~01000000000000abcd",
                                      "title": "Late one", "company": "Acme"}]},
                      headers=ali)
    assert more.status_code == 200, "an open cycle keeps taking jobs after a build"

    assert client.post(f"/api/batches/{batch_id}/close", headers=admin).json()["status"] == "computed"
    late = client.put(f"/api/batches/{batch_id}/profiles/{one}/entries",
                      json={"rows": [{"url": "https://www.upwork.com/jobs/~01000000000000abce"}]},
                      headers=ali)
    assert late.status_code == 400


def test_a_rebuild_keeps_work_people_have_marked(client, admin, new_profile):
    """The timer rebuilds behind people's backs. A job someone marked applied
    is their record of this cycle and must survive that."""
    one = new_profile("ali@example.com")
    two = new_profile("sara@example.com")
    ali = token(client, "ali@example.com")
    batch_id = open_cycle(client, admin, "Marks survive")
    hand_in(client, batch_id, ali, one, sheet(range(9500, 9510)))
    hand_in(client, batch_id, token(client, "sara@example.com"), two, sheet(range(9510, 9520)))
    client.post(f"/api/batches/{batch_id}/compute", headers=admin)

    mine = listing(client, batch_id, one, ali)
    assert len(mine) == 10
    client.patch(f"/api/assignments/{mine[0]['id']}", json={"status": "applied"}, headers=ali)
    client.patch(f"/api/assignments/{mine[1]['id']}", json={"status": "skipped"}, headers=ali)

    client.post(f"/api/batches/{batch_id}/compute", headers=admin)
    after = listing(client, batch_id, one, ali)
    marks = {job["id"]: job["status"] for job in after}
    assert marks.get(mine[0]["id"]) == "applied", "an applied job disappeared on rebuild"
    assert marks.get(mine[1]["id"]) == "skipped", "a skipped job disappeared on rebuild"
    assert len(after) == 10, "the rest of the list is unchanged"


def test_the_timer_builds_an_open_cycle_on_its_own(client, admin, new_profile):
    """No manager, no button — run_due_builds is what the background loop calls."""
    from app.main import run_due_builds

    one = new_profile("ali@example.com")
    two = new_profile("sara@example.com")
    ali = token(client, "ali@example.com")
    batch_id = open_cycle(client, admin, "Hands off", auto_build_minutes=10)
    hand_in(client, batch_id, ali, one, sheet(range(9600, 9610)))
    hand_in(client, batch_id, token(client, "sara@example.com"), two, sheet(range(9610, 9620)))

    assert listing(client, batch_id, one, ali) == [], "nothing built yet"
    assert batch_id in run_due_builds(), "a never-built open cycle is due immediately"
    assert len(listing(client, batch_id, one, ali)) == 10

    # Just built, so the timer leaves it alone until the interval is up.
    assert batch_id not in run_due_builds()

    # And it stops entirely once the cycle is closed.
    client.post(f"/api/batches/{batch_id}/close", headers=admin)
    from app.main import SessionLocal
    from app.models import Batch
    db = SessionLocal()
    try:
        batch = db.get(Batch, batch_id)
        batch.last_built_at = None
        db.commit()
    finally:
        db.close()
    assert batch_id not in run_due_builds(), "a closed cycle must not be rebuilt"


def test_turning_the_timer_off_leaves_it_to_the_manager(client, admin, new_profile):
    from app.main import run_due_builds

    one = new_profile("ali@example.com")
    two = new_profile("sara@example.com")
    ali = token(client, "ali@example.com")
    batch_id = open_cycle(client, admin, "Manual only", auto_build_minutes=0)
    hand_in(client, batch_id, ali, one, sheet(range(9700, 9710)))
    hand_in(client, batch_id, token(client, "sara@example.com"), two, sheet(range(9710, 9720)))

    assert batch_id not in run_due_builds()
    assert listing(client, batch_id, one, ali) == []
    client.post(f"/api/batches/{batch_id}/compute", headers=admin)
    assert len(listing(client, batch_id, one, ali)) == 10


def test_profile_names_must_be_unique(client, admin):
    clash = client.post("/api/profiles", json={"name": "khuram", "headline": "AI"},
                        headers=admin)
    assert clash.status_code == 409


def test_a_bd_cannot_act_as_someone_elses_profile(client, admin, seeded):
    batch_id = open_cycle(client, admin, "Boundary")
    denied = client.post(f"/api/batches/{batch_id}/uploads",
                         data={"profile_id": seeded["Zahid"]},
                         files={"file": ("s.csv", sheet(range(10)), "text/csv")},
                         headers=token(client, "ali@example.com"))
    assert denied.status_code == 403


def test_uploaded_links_cannot_carry_a_script(client, admin, new_profile):
    one = new_profile("ali@example.com")
    two = new_profile("sara@example.com")
    batch_id = open_cycle(client, admin, "Hostile")
    evil = ("Job Title,Company,Job URL\n"
            "Designer,Acme,javascript:alert(document.domain)\n").encode()
    hand_in(client, batch_id, token(client, "ali@example.com"), one, evil)
    hand_in(client, batch_id, token(client, "sara@example.com"), two, sheet(range(6000, 6005)))
    assert client.post(f"/api/batches/{batch_id}/compute", headers=admin).status_code == 200

    for profile_id, email in ((one, "ali@example.com"), (two, "sara@example.com")):
        jobs = listing(client, batch_id, profile_id, token(client, email))
        assert not any((j["url"] or "").lower().startswith("javascript:") for j in jobs)


def test_exported_cells_cannot_be_formulas(client, admin, new_profile):
    one = new_profile("ali@example.com")
    two = new_profile("sara@example.com")
    batch_id = open_cycle(client, admin, "Formula")
    nasty = ("Job Title,Company,Job URL\n"
             "\"=cmd|'/c calc'!A1\",Acme,https://www.upwork.com/jobs/~01000000000000f1f1\n").encode()
    hand_in(client, batch_id, token(client, "ali@example.com"), one, nasty)
    hand_in(client, batch_id, token(client, "sara@example.com"), two, sheet(range(7000, 7005)))
    assert client.post(f"/api/batches/{batch_id}/compute", headers=admin).status_code == 200

    from io import BytesIO

    from openpyxl import load_workbook
    for profile_id, email in ((one, "ali@example.com"), (two, "sara@example.com")):
        response = client.get(f"/api/batches/{batch_id}/profiles/{profile_id}/sheet.xlsx",
                              headers=token(client, email))
        if response.status_code != 200:
            continue
        book = load_workbook(BytesIO(response.content))
        for page in book.worksheets:
            for row in page.iter_rows(min_row=2, values_only=True):
                for value in row:
                    assert not (isinstance(value, str)
                                and value.startswith(("=", "+", "@"))), value


def test_report_counts_only_its_own_cycle(client, admin, new_profile):
    one = new_profile("ali@example.com")
    two = new_profile("sara@example.com")

    first = open_cycle(client, admin, "Scoped A")
    hand_in(client, first, token(client, "ali@example.com"), one, sheet(range(8000, 8010)))
    hand_in(client, first, token(client, "sara@example.com"), two, sheet(range(8005, 8015)))
    computed = client.post(f"/api/batches/{first}/compute", headers=admin).json()
    before = client.get(f"/api/batches/{first}/report", headers=admin).json()

    later = open_cycle(client, admin, "Scoped B")
    hand_in(client, later, token(client, "ali@example.com"), one, sheet(range(8100, 8110)))
    hand_in(client, later, token(client, "sara@example.com"), two, sheet(range(8105, 8115)))
    client.post(f"/api/batches/{later}/compute", headers=admin)

    after = client.get(f"/api/batches/{first}/report", headers=admin).json()
    assert after["matrix"] == before["matrix"], "a later cycle changed an earlier report"
    assert after["matrix"]["rows"] == computed["matrix"]["rows"]


def test_bd_cannot_compute(client):
    headers = token(client, "ali@example.com")
    assert client.post("/api/batches", json={"name": "x"}, headers=headers).status_code == 403


def test_anonymous_is_blocked(client):
    assert client.get("/api/batches").status_code == 401
