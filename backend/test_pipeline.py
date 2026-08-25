"""The record, the ladder, and the work that comes out of a conversation.

    pytest test_pipeline.py -v

What is pinned down here:

  * every job a profile ever applied to stays findable, all-time and by search,
    because a client's reply arrives long after the cycle that earned it closed;
  * an interview can be started from that record with no time agreed — a draft —
    and counts towards nothing until somebody puts a time on it;
  * the stage ladder, and the fact that where a conversation died is carried
    rather than flattened into one interviews-to-offers percentage;
  * assessments: set by either side, done by the developer, and late when the
    deadline has passed and nobody has submitted;
  * a profile that has handed in nothing still receives the pool when it sells
    the same skills as somebody who has.

Shares a database with the rest of the suite — pytest imports app.main once, so
whichever module runs first decides where the data lives. Every account and
profile minted here carries a name no other module uses.
"""
import datetime as dt
import io
import itertools

import pytest
from fastapi.testclient import TestClient

TEAM = [("pl-boss@pipeline.example.com", "Pipe Boss", "admin"),
        ("pl-bd@pipeline.example.com", "Pia Dee", "bd"),
        ("pl-bd2@pipeline.example.com", "Fay Zann", "bd"),
        ("pl-dev@pipeline.example.com", "Dev Pipe", "dev")]

PASSWORD = "pipepass12345"
_serial = itertools.count(1)


@pytest.fixture(scope="module")
def client(tmp_path_factory):
    import os
    os.environ.setdefault("DATABASE_URL",
                          f"sqlite:///{tmp_path_factory.mktemp('db')}/pipe.db")
    from app.main import SessionLocal, app, engine, hash_password  # noqa: E402
    from app.models import Base, User  # noqa: E402

    Base.metadata.create_all(engine)
    db = SessionLocal()
    for email, name, role in TEAM:
        if not db.query(User).filter(User.email == email).first():
            db.add(User(email=email, name=name, dashboard_visible=True,
                        password_hash=hash_password(PASSWORD), role=role))
    db.commit()
    db.close()
    return TestClient(app)


def token(client, email):
    response = client.post("/api/auth/login", json={"email": email, "password": PASSWORD})
    assert response.status_code == 200, response.text
    return {"Authorization": f"Bearer {response.json()['token']}"}


@pytest.fixture(scope="module")
def admin(client):
    return token(client, "pl-boss@pipeline.example.com")


@pytest.fixture(scope="module")
def bd(client):
    return token(client, "pl-bd@pipeline.example.com")


@pytest.fixture(scope="module")
def other_bd(client):
    return token(client, "pl-bd2@pipeline.example.com")


@pytest.fixture(scope="module")
def dev(client):
    return token(client, "pl-dev@pipeline.example.com")


@pytest.fixture
def people(client, admin):
    return {u["email"]: u["id"] for u in client.get("/api/users", headers=admin).json()}


@pytest.fixture
def make_profile(client, admin, people):
    def make(bd_email="pl-bd@pipeline.example.com",
             dev_email="pl-dev@pipeline.example.com", skills="", **extra):
        body = {"name": f"Pipe{next(_serial)}", "headline": "AI Engineer",
                "platform": "Upwork", "user_id": people[bd_email],
                "dev_user_id": people[dev_email] if dev_email else None,
                "skills": skills}
        body.update(extra)
        response = client.post("/api/profiles", json=body, headers=admin)
        assert response.status_code == 201, response.text
        return response.json()
    return make


def when(days: int, clock: str) -> str:
    from app.models import working_today
    return f"{(working_today() + dt.timedelta(days=days)).isoformat()}T{clock}"


def sheet(job_ids, client_name="Northwind Digital", title="RAG Developer"):
    out = io.StringIO()
    out.write("Job Title,Company,Job URL,Job description link\n")
    for n in job_ids:
        out.write(f"{title} {n},{client_name},"
                  f"https://www.upwork.com/jobs/~03{n:016x},"
                  f"https://example.com/jd/{n}\n")
    return out.getvalue().encode()


def cycle(client, admin, name, **extra):
    body = {"name": name, "mode": "cover", "quota": 500, "auto_build_minutes": 0}
    body.update(extra)
    response = client.post("/api/batches", json=body, headers=admin)
    assert response.status_code == 201, response.text
    return response.json()["id"]


def hand_in(client, batch_id, headers, profile_id, payload):
    response = client.post(f"/api/batches/{batch_id}/uploads",
                           data={"profile_id": profile_id},
                           files={"file": ("s.csv", payload, "text/csv")}, headers=headers)
    assert response.status_code == 201, response.text
    return response.json()


@pytest.fixture
def applied(client, admin, bd, make_profile):
    """A profile with a real application history behind it."""
    mine = make_profile()
    theirs = make_profile()
    batch = cycle(client, admin, f"Rec{next(_serial)}")
    hand_in(client, batch, bd, mine["id"], sheet(range(700, 706)))
    hand_in(client, batch, bd, theirs["id"], sheet(range(704, 710), "Sable Analytics"))
    client.post(f"/api/batches/{batch}/compute", headers=admin)
    return {"profile": mine, "other": theirs, "batch": batch}


# --------------------------------------------------------------------------- #
# The record
# --------------------------------------------------------------------------- #

def test_the_record_holds_everything_applied_for(client, bd, applied):
    """Not scoped to a cycle. A reply arrives three weeks after the application
    that earned it, by which time that cycle is closed and off the screens."""
    record = client.get("/api/jobs", headers=bd).json()
    assert record["total"] >= 12
    row = record["rows"][0]
    assert {"job_id", "profile", "title", "company", "url", "description_url"} <= set(row)


def test_the_record_is_searched_by_anything_pasted_out_of_an_email(client, bd, applied):
    """A person pasting out of a client's email has no idea which field the
    thing they copied lives in, so all of them are searched."""
    # Scoped to the one profile. This fixture mints fresh profiles per test
    # against the same postings, so the workspace-wide count grows as the module
    # runs and only a scoped one is a number worth asserting on.
    mine = applied["other"]["id"]
    by_client = client.get(f"/api/jobs?q=sable&profile_id={mine}", headers=bd).json()
    # Four of its six, not all six. Jobs 704 and 705 are on both sheets — the
    # same posting found twice — and a job is one row: the first client name
    # recorded for it is the one it keeps. That is the fingerprint doing its
    # job, showing up here as two of Sable's six filed under Northwind.
    assert by_client["total"] == 4
    assert all(r["company"] == "Sable Analytics" for r in by_client["rows"])

    # Mid-value, not just leading — "Northwind" inside "The Northwind Group".
    by_title = client.get(f"/api/jobs?q=Developer+704&profile_id={mine}", headers=bd).json()
    assert by_title["total"] == 1

    assert client.get("/api/jobs?q=nothingmatchesthis", headers=bd).json()["total"] == 0


def test_the_record_stops_at_the_profiles_you_run(client, other_bd, applied):
    """A BD who runs neither profile sees none of it."""
    assert client.get("/api/jobs", headers=other_bd).json()["total"] == 0


def test_the_record_pages(client, bd, applied):
    mine = applied["profile"]["id"]
    first = client.get(f"/api/jobs?limit=5&profile_id={mine}", headers=bd).json()
    second = client.get(f"/api/jobs?limit=5&offset=5&profile_id={mine}", headers=bd).json()
    assert len(first["rows"]) == 5
    assert first["total"] == second["total"] == 6
    # Keyed on the pair, not the job. One posting legitimately appears twice
    # when two profiles both applied to it — they are two candidates, and the
    # record is of applications rather than of postings.
    def keys(page):
        return {(r["job_id"], r["profile_id"]) for r in page["rows"]}
    assert keys(first) & keys(second) == set()


def test_the_description_link_survives_the_sheet(client, bd, applied):
    """The apply link dies when the posting expires and takes the wording with
    it. The description link is what a BD still has three weeks later."""
    row = client.get("/api/jobs?q=sable", headers=bd).json()["rows"][0]
    assert row["description_url"].startswith("https://example.com/jd/")


# --------------------------------------------------------------------------- #
# Starting an interview from the record
# --------------------------------------------------------------------------- #

def test_an_interview_started_from_a_job_carries_the_job_across(client, bd, applied):
    """The point of the record. The title, the client and both links come over
    instead of being typed again out of an email, wrong."""
    job = client.get("/api/jobs?q=sable", headers=bd).json()["rows"][0]
    made = client.post("/api/interviews",
                       json={"profile_id": applied["other"]["id"], "job_id": job["job_id"]},
                       headers=bd)
    assert made.status_code == 201, made.text
    row = made.json()
    assert row["status"] == "draft"
    assert row["is_draft"] is True
    assert row["client"] == "Sable Analytics"      # filled from the job
    assert row["role"] == job["title"]
    assert row["job"]["description_url"] == job["description_url"]


def test_a_draft_counts_towards_nothing_until_it_has_a_time(client, bd, applied):
    """No time agreed means no conversation yet. Counting it would put a client
    who has not answered into a rate that says they did."""
    profile = applied["profile"]["id"]
    before = client.get(f"/api/interviews?profile_id={profile}", headers=bd).json()

    draft = client.post("/api/interviews", json={"profile_id": profile,
                                                 "client": "Ironvale"},
                        headers=bd).json()
    after = client.get(f"/api/interviews?profile_id={profile}", headers=bd).json()

    assert after["counts"]["awaiting_time"] == before["counts"]["awaiting_time"] + 1
    assert after["counts"]["scheduled"] == before["counts"]["scheduled"]
    assert after["funnel"]["interviews"] == before["funnel"]["interviews"]
    # Its own list — a reply nobody has answered is a thing to chase.
    assert draft["id"] in {r["id"] for r in after["awaiting_time"]}
    assert draft["id"] not in {r["id"] for r in after["today"] + after["upcoming"]}


def test_putting_a_time_on_a_draft_is_what_books_it(client, bd, applied):
    """Agreeing the time *is* the confirmation. A second button to say so is a
    button somebody forgets, leaving a real interview counted nowhere."""
    profile = applied["profile"]["id"]
    draft = client.post("/api/interviews", json={"profile_id": profile,
                                                 "client": "Ironvale Systems"},
                        headers=bd).json()

    booked = client.patch(f"/api/interviews/{draft['id']}",
                          json={"scheduled_at": when(2, "14:00")}, headers=bd)
    assert booked.status_code == 200, booked.text
    assert booked.json()["status"] == "scheduled"
    assert booked.json()["when"]["time"] == "14:00"


def test_a_draft_cannot_be_reported_on(client, bd, applied):
    """It has no time on it, so it has not happened. An outcome here would be
    a call nobody had, counted in a rate."""
    draft = client.post("/api/interviews",
                        json={"profile_id": applied["profile"]["id"], "client": "Ghost Co"},
                        headers=bd).json()
    refused = client.patch(f"/api/interviews/{draft['id']}",
                           json={"outcome": "offer"}, headers=bd)
    assert refused.status_code == 400
    assert "has not happened" in refused.json()["detail"]


# --------------------------------------------------------------------------- #
# The ladder
# --------------------------------------------------------------------------- #

def test_a_conversation_climbs_the_ladder(client, bd, dev, applied):
    """Where a conversation died matters more than that it died — a team losing
    everybody at technical has a different problem from one losing them at
    final, and a single percentage cannot tell those apart."""
    profile = applied["profile"]["id"]
    first = client.post("/api/interviews",
                        json={"profile_id": profile, "scheduled_at": when(-3, "10:00"),
                              "client": "Larkspur Data", "stage": "screening"},
                        headers=bd).json()
    assert first["stage"] == "screening"

    client.patch(f"/api/interviews/{first['id']}", json={"outcome": "passed"}, headers=dev)

    second = client.post("/api/interviews",
                         json={"profile_id": profile, "scheduled_at": when(-1, "10:00"),
                               "client": "Larkspur Data", "stage": "technical",
                               "job_id": first["job_id"]},
                         headers=bd).json()
    client.patch(f"/api/interviews/{second['id']}", json={"outcome": "rejected"}, headers=dev)

    stages = {row["stage"]: row for row in
              client.get(f"/api/interviews?profile_id={profile}",
                         headers=bd).json()["funnel"]["by_stage"]}
    assert stages["screening"]["cleared"] >= 1
    assert stages["technical"]["lost"] >= 1
    assert [s["stage"] for s in
            client.get(f"/api/interviews?profile_id={profile}",
                       headers=bd).json()["funnel"]["by_stage"]] == \
        ["screening", "technical", "assessment", "final", "offer"]


def test_a_stage_has_to_be_one_of_the_rungs(client, bd, applied):
    refused = client.post("/api/interviews",
                          json={"profile_id": applied["profile"]["id"],
                                "scheduled_at": when(1, "09:00"), "stage": "vibes"},
                          headers=bd)
    assert refused.status_code == 400


# --------------------------------------------------------------------------- #
# Assessments
# --------------------------------------------------------------------------- #

def test_the_bd_sets_it_and_the_developer_does_it(client, bd, dev, applied):
    """The same split as an interview: the BD has the client's email, the
    developer has the work."""
    profile = applied["profile"]["id"]
    sitting = client.post("/api/interviews",
                          json={"profile_id": profile, "scheduled_at": when(-2, "11:00"),
                                "client": "Copperline Media", "stage": "technical"},
                          headers=bd).json()

    made = client.post("/api/assessments",
                       json={"profile_id": profile, "interview_id": sitting["id"],
                             "brief": "Build a small RAG pipeline.",
                             "link": "https://example.com/take-home",
                             "due_at": when(3, "17:00")},
                       headers=bd)
    assert made.status_code == 201, made.text
    row = made.json()
    assert row["status"] == "sent"
    assert row["client"] == "Copperline Media"     # filled from the interview
    assert row["interview"]["stage"] == "technical"
    assert row["set_by"] == "Pia Dee"

    done = client.patch(f"/api/assessments/{row['id']}",
                        json={"status": "submitted",
                              "submission_url": "https://github.com/example/take-home",
                              "notes": "Went with hybrid search. Took about four hours."},
                        headers=dev)
    assert done.status_code == 200, done.text
    assert done.json()["submitted"] is not None
    assert done.json()["is_open"] is False

    # And the BD reads it on their own screen.
    seen = client.get(f"/api/assessments?profile_id={profile}", headers=bd).json()
    mine = next(a for a in seen["rows"] if a["id"] == row["id"])
    assert mine["notes"].startswith("Went with hybrid search")
    assert mine["updated_by"] == "Dev Pipe"


def test_an_assessment_nobody_submitted_goes_overdue(client, bd, applied):
    """A deadline nobody is watching is the same as no deadline, and a missed
    take-home costs the interview that earned it."""
    profile = applied["profile"]["id"]
    late = client.post("/api/assessments",
                       json={"profile_id": profile, "title": "Late one",
                             "due_at": when(-2, "17:00")}, headers=bd).json()
    assert late["overdue"] is True

    figures = client.get(f"/api/assessments?profile_id={profile}", headers=bd).json()
    assert figures["counts"]["overdue"] >= 1
    assert late["id"] in {a["id"] for a in figures["open"]}


def test_a_submitted_assessment_is_not_late_however_late_it_was(client, bd, dev, applied):
    """Only an open one can be chased. One that went in after the deadline is a
    thing that happened, not a thing to do."""
    profile = applied["profile"]["id"]
    row = client.post("/api/assessments",
                      json={"profile_id": profile, "title": "Handed in late",
                            "due_at": when(-4, "17:00")}, headers=bd).json()
    assert row["overdue"] is True

    after = client.patch(f"/api/assessments/{row['id']}",
                         json={"status": "submitted"}, headers=dev).json()
    assert after["overdue"] is False


def test_an_assessment_with_no_deadline_is_not_late(client, bd, applied):
    """No deadline is the commonest answer after "next Friday". Inventing one
    puts a red flag on a screen nobody set."""
    row = client.post("/api/assessments",
                      json={"profile_id": applied["profile"]["id"], "title": "Open ended"},
                      headers=bd).json()
    assert row["due"] is None
    assert row["overdue"] is False
    assert row["is_open"] is True


def test_a_stranger_cannot_reach_the_assessments(client, other_bd, bd, applied):
    profile = applied["profile"]["id"]
    client.post("/api/assessments", json={"profile_id": profile, "title": "Private"},
                headers=bd)
    refused = client.get(f"/api/assessments?profile_id={profile}", headers=other_bd)
    assert refused.status_code == 403


def test_an_assessment_cannot_be_hung_on_another_profiles_interview(
        client, bd, applied, make_profile):
    elsewhere = make_profile()
    sitting = client.post("/api/interviews",
                          json={"profile_id": elsewhere["id"],
                                "scheduled_at": when(1, "10:00")}, headers=bd).json()
    refused = client.post("/api/assessments",
                          json={"profile_id": applied["profile"]["id"],
                                "interview_id": sitting["id"]}, headers=bd)
    assert refused.status_code == 400


# --------------------------------------------------------------------------- #
# Joining a cycle you handed nothing into
# --------------------------------------------------------------------------- #

def test_a_profile_selling_the_same_skills_still_gets_the_pool(
        client, admin, bd, make_profile):
    """The profile with the most spare capacity should not be the one handed no
    work just because it has not logged anything yet."""
    one = make_profile(skills="Python, RAG, LangChain")
    two = make_profile(skills="Python, PyTorch")
    idle = make_profile(skills="RAG, LLM")            # hands in nothing
    unrelated = make_profile(skills="Figma, Illustrator")

    batch = cycle(client, admin, f"Join{next(_serial)}")
    hand_in(client, batch, bd, one["id"], sheet(range(800, 806)))
    hand_in(client, batch, bd, two["id"], sheet(range(806, 812), "Verdant Labs"))
    data = client.post(f"/api/batches/{batch}/compute", headers=admin).json()

    given = {p["id"]: p["assigned"] for p in data["participants"]}
    assert given.get(idle["id"], 0) == 12, "a profile selling the same skills gets the pool"
    assert unrelated["id"] not in given, "a designer is not in an AI cycle"
    # At least the one this test made. The suite shares a database, so other
    # modules' profiles can legitimately match too — which is why every precise
    # assertion here is about a profile this test minted itself.
    assert data["report"].get("Profiles pulled in on skills") >= 1


def test_a_profile_with_no_skills_recorded_is_never_pulled_in(
        client, admin, bd, make_profile):
    """An empty field is not a match. It is a profile nobody has finished
    setting up, and handing it a stranger's pool would be a guess."""
    one = make_profile(skills="Rust, WebAssembly")
    two = make_profile(skills="Rust, Tokio")
    blank = make_profile(skills="")

    batch = cycle(client, admin, f"Blank{next(_serial)}")
    hand_in(client, batch, bd, one["id"], sheet(range(900, 904)))
    hand_in(client, batch, bd, two["id"], sheet(range(904, 908), "Talloak"))
    data = client.post(f"/api/batches/{batch}/compute", headers=admin).json()

    assert blank["id"] not in {p["id"] for p in data["participants"]}


def test_shared_tooling_is_not_the_same_market(client, admin, bd, make_profile):
    """Everybody runs Postgres in Docker on AWS. Matching on that is how a
    front-end profile lands in an AI cycle, so shared infrastructure cannot
    carry a match on its own — a shared language or framework can."""
    one = make_profile(skills="Python, RAG, LangChain, Postgres, Docker")
    two = make_profile(skills="Python, PyTorch, MLOps, Postgres")
    frontend = make_profile(skills="React, Stripe, Postgres, Docker, AWS")

    batch = cycle(client, admin, f"Infra{next(_serial)}")
    hand_in(client, batch, bd, one["id"], sheet(range(970, 974)))
    hand_in(client, batch, bd, two["id"], sheet(range(974, 978), "Larkspur"))
    data = client.post(f"/api/batches/{batch}/compute", headers=admin).json()

    joined = {p["id"] for p in data["participants"]}
    assert frontend["id"] not in joined, "Postgres and Docker in common is not a market"


def test_a_late_joiner_is_given_jobs_not_given_away(client, admin, bd, make_profile):
    """They contributed nothing to the pool, so nothing of theirs is in it —
    and their own history still blocks anything they have already applied to."""
    one = make_profile(skills="Elixir, Phoenix")
    two = make_profile(skills="Elixir, OTP")
    idle = make_profile(skills="Phoenix, Elixir")

    first = cycle(client, admin, f"Late1{next(_serial)}")
    hand_in(client, first, bd, idle["id"], sheet(range(950, 953)))
    hand_in(client, first, bd, one["id"], sheet(range(953, 956), "Harbourstone"))
    client.post(f"/api/batches/{first}/compute", headers=admin)

    second = cycle(client, admin, f"Late2{next(_serial)}")
    hand_in(client, second, bd, one["id"], sheet(range(950, 956)))
    hand_in(client, second, bd, two["id"], sheet(range(956, 960), "Orchard"))
    data = client.post(f"/api/batches/{second}/compute", headers=admin).json()

    given = {p["id"]: p["assigned"] for p in data["participants"]}
    # 950-952 are its own history and never come back; 953-959 are new to it.
    assert given[idle["id"]] == 7


# --------------------------------------------------------------------------- #
# Rounds — one conversation, several sittings
# --------------------------------------------------------------------------- #

def sat(client, headers, profile_id, days=-3, clock="10:00", **extra):
    """An interview that has already been and gone, unless days says otherwise."""
    body = {"profile_id": profile_id, "scheduled_at": when(days, clock),
            "client": "Larkspur Data", "role": "RAG Developer"}
    body.update(extra)
    response = client.post("/api/interviews", json=body, headers=headers)
    assert response.status_code == 201, response.text
    return response.json()


def test_booking_the_next_round_carries_the_conversation_across(client, bd, applied):
    """The gap this closes loses work.

    A screening call goes well and the second round then has to be retyped from
    memory — same client, same role, same posting — so it usually is not typed
    for a week, and the client is left waiting on a team that thinks it is
    winning. One press instead, and nothing is retyped.
    """
    profile = applied["profile"]["id"]
    record = client.get(f"/api/jobs?profile_id={profile}", headers=bd).json()["rows"][0]
    first = sat(client, bd, profile, stage="screening", job_id=record["job_id"])

    made = client.post(f"/api/interviews/{first['id']}/next-round", json={}, headers=bd)
    assert made.status_code == 201, made.text
    second = made.json()

    assert second["client"] == "Larkspur Data"
    assert second["role"] == "RAG Developer"
    assert second["job_id"] == record["job_id"]
    assert second["stage"] == "technical"          # one rung up, guessed for them
    assert second["previous_id"] == first["id"]
    # No time was agreed, so it waits rather than inventing an appointment.
    assert second["status"] == "draft"
    assert second["round"] == 2 and second["rounds"] == 2

    # And booking it is what says the round before was cleared.
    assert second["previous"]["outcome"] == "passed"
    assert second["previous"]["status"] == "done"


def test_booking_the_next_round_does_not_overwrite_a_recorded_outcome(client, bd, dev,
                                                                      applied):
    """An offer somebody recorded is theirs. A scheduling action must not
    quietly demote it to `passed`."""
    profile = applied["profile"]["id"]
    first = sat(client, bd, profile, stage="final")
    client.patch(f"/api/interviews/{first['id']}", json={"outcome": "offer"}, headers=dev)

    made = client.post(f"/api/interviews/{first['id']}/next-round", json={}, headers=bd)
    assert made.status_code == 201, made.text
    assert made.json()["previous"]["outcome"] == "offer"


def test_a_round_that_has_not_happened_yet_is_not_marked_done(client, bd, applied):
    """A client saying up front that there will be two rounds is not a report
    on the first one, and a call next Tuesday must not read as finished."""
    profile = applied["profile"]["id"]
    ahead = sat(client, bd, profile, days=4, clock="15:00", stage="screening")

    made = client.post(f"/api/interviews/{ahead['id']}/next-round", json={}, headers=bd)
    assert made.status_code == 201, made.text
    assert made.json()["previous"]["status"] == "scheduled"
    assert made.json()["previous"]["outcome"] == "pending"


def test_nothing_follows_a_round_that_was_a_no(client, bd, dev, applied):
    profile = applied["profile"]["id"]
    first = sat(client, bd, profile)
    client.patch(f"/api/interviews/{first['id']}", json={"outcome": "rejected"}, headers=dev)

    refused = client.post(f"/api/interviews/{first['id']}/next-round", json={}, headers=bd)
    assert refused.status_code == 400
    assert "no" in refused.json()["detail"].lower()


def test_a_reply_with_no_time_on_it_cannot_have_a_second_round(client, bd, applied):
    """A draft has not happened. There is nothing for a next round to follow."""
    profile = applied["profile"]["id"]
    draft = client.post("/api/interviews",
                        json={"profile_id": profile, "scheduled_at": "",
                              "client": "Larkspur Data"}, headers=bd).json()
    assert draft["status"] == "draft"

    refused = client.post(f"/api/interviews/{draft['id']}/next-round", json={}, headers=bd)
    assert refused.status_code == 400


def test_a_chain_of_rounds_reads_as_one_conversation(client, bd, applied):
    """Four rounds with one client is a different fact from four clients who
    each said no after one call, and in a flat list they look identical."""
    profile = applied["profile"]["id"]
    first = sat(client, bd, profile, stage="screening")
    second = client.post(f"/api/interviews/{first['id']}/next-round",
                         json={"scheduled_at": when(-2, "11:00")}, headers=bd).json()
    third = client.post(f"/api/interviews/{second['id']}/next-round",
                        json={"scheduled_at": when(-1, "11:00")}, headers=bd).json()

    rows = {row["id"]: row for row in
            client.get(f"/api/interviews?profile_id={profile}", headers=bd).json()["rows"]}
    assert [rows[first["id"]]["round"], rows[second["id"]]["round"],
            rows[third["id"]]["round"]] == [1, 2, 3]
    assert rows[third["id"]]["rounds"] == 3
    # Read from the middle, it still knows what came before it.
    assert rows[second["id"]]["follows"]["stage"] == "screening"
    assert rows[first["id"]]["follows"] is None


def test_a_chain_cannot_cross_two_profiles(client, bd, applied):
    """A chain of rounds is one client talking to one identity. Linking across
    two would make a stranger's conversation part of this profile's story."""
    first = sat(client, bd, applied["profile"]["id"])
    refused = client.post("/api/interviews",
                          json={"profile_id": applied["other"]["id"],
                                "scheduled_at": when(1, "09:00"),
                                "previous_id": first["id"]}, headers=bd)
    assert refused.status_code == 400


def test_a_cleared_round_with_nothing_after_it_is_called_out(client, bd, dev, applied):
    """The quietest way this product loses work: the client said yes, nobody
    booked the next round, and on every other screen it reads as a success."""
    profile = applied["profile"]["id"]
    first = sat(client, bd, profile, stage="screening")
    client.patch(f"/api/interviews/{first['id']}", json={"outcome": "passed"}, headers=dev)

    diary = client.get(f"/api/interviews?profile_id={profile}", headers=bd).json()
    assert diary["counts"]["stalled"] == 1
    assert [row["id"] for row in diary["stalled"]] == [first["id"]]

    # Booking what follows is what clears it.
    client.post(f"/api/interviews/{first['id']}/next-round", json={}, headers=bd)
    after = client.get(f"/api/interviews?profile_id={profile}", headers=bd).json()
    assert after["counts"]["stalled"] == 0


def test_a_hire_is_not_a_stalled_conversation(client, bd, dev, applied):
    """Nothing follows a hire. It is the end of the ladder, not a gap in it."""
    profile = applied["profile"]["id"]
    row = sat(client, bd, profile, stage="offer")
    client.patch(f"/api/interviews/{row['id']}", json={"outcome": "hired"}, headers=dev)
    diary = client.get(f"/api/interviews?profile_id={profile}", headers=bd).json()
    assert diary["counts"]["stalled"] == 0


def test_removing_a_round_leaves_what_followed_it_standing(client, bd, applied):
    """The link is a convenience. Losing it costs a breadcrumb; taking a real
    second round away with a mistyped first one would cost the work."""
    profile = applied["profile"]["id"]
    first = sat(client, bd, profile)
    second = client.post(f"/api/interviews/{first['id']}/next-round",
                         json={"scheduled_at": when(-1, "12:00")}, headers=bd).json()

    assert client.delete(f"/api/interviews/{first['id']}", headers=bd).status_code == 200

    rows = {row["id"]: row for row in
            client.get(f"/api/interviews?profile_id={profile}", headers=bd).json()["rows"]}
    assert second["id"] in rows
    assert rows[second["id"]]["previous_id"] is None
    assert rows[second["id"]]["round"] == 1


def test_the_developer_can_book_the_next_round_too(client, bd, dev, applied):
    """They are usually the first to know it is wanted — the client said so in
    the room — and making them ask somebody else to type it in loses a week."""
    profile = applied["profile"]["id"]
    first = sat(client, bd, profile)
    made = client.post(f"/api/interviews/{first['id']}/next-round", json={}, headers=dev)
    assert made.status_code == 201, made.text
    assert made.json()["previous_id"] == first["id"]


def test_a_stranger_cannot_book_a_round_onto_someone_elses_conversation(
        client, bd, other_bd, applied):
    first = sat(client, bd, applied["profile"]["id"])
    refused = client.post(f"/api/interviews/{first['id']}/next-round", json={},
                          headers=other_bd)
    assert refused.status_code == 403


# --------------------------------------------------------------------------- #
# Take-homes, everywhere they have to be seen
# --------------------------------------------------------------------------- #

def test_a_take_home_shows_on_the_interview_it_came_out_of(client, bd, applied):
    """When a client has sent a test, the test *is* the state of that
    conversation. A diary that cannot say so sends people to a second screen to
    find out whether they are waiting on the client or on their own developer.
    """
    profile = applied["profile"]["id"]
    sitting = sat(client, bd, profile, stage="technical")
    client.post("/api/assessments",
                json={"profile_id": profile, "interview_id": sitting["id"],
                      "title": "Small RAG pipeline", "due_at": when(2, "17:00")},
                headers=bd)

    rows = {row["id"]: row for row in
            client.get(f"/api/interviews?profile_id={profile}", headers=bd).json()["rows"]}
    attached = rows[sitting["id"]]["assessments"]
    assert [a["title"] for a in attached] == ["Small RAG pipeline"]


def test_an_overdue_take_home_reaches_every_dashboard(client, admin, bd, dev, applied):
    """A missed deadline is the one thing here that goes wrong silently. Every
    other figure understates itself when nobody touches it; this one changes
    nothing on any screen until the client's next email — so it has to be on
    all of them.
    """
    profile = applied["profile"]["id"]
    client.post("/api/assessments",
                json={"profile_id": profile, "title": "Late take-home",
                      "due_at": when(-2, "17:00")},
                headers=bd)

    mine = client.get("/api/dashboard/me", headers=bd).json()
    assert mine["assessments"]["counts"]["overdue"] >= 1

    org = client.get("/api/dashboard/overview", headers=admin).json()
    assert org["assessments"]["counts"]["overdue"] >= 1

    desk = client.get("/api/dashboard/dev", headers=dev).json()
    assert desk["assessments"]["counts"]["overdue"] >= 1

    detail = client.get(f"/api/dashboard/profiles/{profile}", headers=bd).json()
    assert detail["assessments"]["counts"]["overdue"] == 1

    # And on the profile's own row, so a board about how much went out cannot
    # stay silent about the test that could lose it.
    row = next(p for p in mine["profiles"] if p["profile_id"] == profile)
    assert row["assessments_open"] == 1
    assert row["assessments_overdue"] == 1


def test_the_developer_board_says_who_owes_a_take_home(client, admin, bd, applied):
    """Free on Thursday is not free when a test is due Friday, and a manager
    reading availability off the diary alone books straight over it."""
    profile = applied["profile"]["id"]
    client.post("/api/assessments",
                json={"profile_id": profile, "title": "Weekend eater",
                      "due_at": when(1, "17:00")},
                headers=bd)

    board = client.get("/api/dashboard/overview", headers=admin).json()["developers"]
    mine = next(row for row in board if row["name"] == "Dev Pipe")
    assert mine["assessments_open"] >= 1


def test_the_pipeline_downloads_as_a_spreadsheet(client, bd, dev, applied):
    """Everything downloadable until now counted what went out, and a team can
    improve all of that without winning a single piece of work. This is the
    other half — the part somebody is asked about at the end of a quarter."""
    from openpyxl import load_workbook

    profile = applied["profile"]["id"]
    first = sat(client, bd, profile, stage="screening")
    client.post(f"/api/interviews/{first['id']}/next-round",
                json={"scheduled_at": when(-1, "10:00")}, headers=bd)
    client.post("/api/assessments",
                json={"profile_id": profile, "title": "Small RAG pipeline",
                      "due_at": when(-1, "17:00")},
                headers=bd)

    got = client.get(f"/api/pipeline.xlsx?profile_id={profile}", headers=bd)
    assert got.status_code == 200, got.text
    book = load_workbook(io.BytesIO(got.content))
    assert book.sheetnames == ["Conversations", "Take-homes"]

    talks = [[c.value for c in row] for row in book["Conversations"].iter_rows(min_row=2)]
    assert any("2 of 2" in (row[5] or "") for row in talks)     # the chain is legible

    tests = [[c.value for c in row] for row in book["Take-homes"].iter_rows(min_row=2)]
    # Said in words: a file outlives the screen that would explain a status.
    assert any("overdue" in (row[5] or "") for row in tests)


def test_a_stranger_cannot_download_someone_elses_pipeline(client, bd, other_bd, applied):
    sat(client, bd, applied["profile"]["id"])
    refused = client.get(f"/api/pipeline.xlsx?profile_id={applied['profile']['id']}",
                         headers=other_bd)
    assert refused.status_code == 403


def test_the_dispatched_sheet_keeps_who_found_each_job(client, admin, bd, applied):
    """The download is what somebody works through on a train, and it should
    not be a worse document than the screen it came from."""
    from openpyxl import load_workbook

    got = client.get(f"/api/batches/{applied['batch']}/profiles/"
                     f"{applied['profile']['id']}/sheet.xlsx", headers=bd)
    assert got.status_code == 200, got.text
    sheet = load_workbook(io.BytesIO(got.content)).worksheets[0]
    headers = [cell.value for cell in sheet[1]]
    assert "Found by" in headers and "Description link" in headers

    column = headers.index("Found by")
    rows = [[c.value for c in row] for row in sheet.iter_rows(min_row=2)]
    assert any(row[column] for row in rows), "nothing said who found these"


def test_nothing_follows_a_round_that_was_called_off(client, bd, applied):
    """A cancelled call never happened, so it cleared nothing. Marking it
    passed on the way to booking a second round would invent a conversation."""
    profile = applied["profile"]["id"]
    first = sat(client, bd, profile)
    client.patch(f"/api/interviews/{first['id']}", json={"status": "cancelled"}, headers=bd)

    refused = client.post(f"/api/interviews/{first['id']}/next-round", json={}, headers=bd)
    assert refused.status_code == 400
    assert "called off" in refused.json()["detail"]

    # And it is still a cancellation afterwards, not a pass.
    rows = {row["id"]: row for row in
            client.get(f"/api/interviews?profile_id={profile}", headers=bd).json()["rows"]}
    assert rows[first["id"]]["status"] == "cancelled"
    assert rows[first["id"]]["outcome"] == "pending"


def test_the_whole_posting_travels_onto_the_interview(client, bd, applied):
    """Six things, and a BD reading a client's reply three weeks later wants
    all of them. Going back to the record for the platform or the day it went
    out is exactly the retyping the record exists to stop."""
    profile = applied["profile"]["id"]
    record = next(r for r in
                  client.get(f"/api/jobs?profile_id={profile}", headers=bd).json()["rows"]
                  if r["url"] and r["description_url"])

    made = client.post("/api/interviews",
                       json={"profile_id": profile, "scheduled_at": "",
                             "job_id": record["job_id"]}, headers=bd)
    assert made.status_code == 201, made.text
    job = made.json()["job"]

    assert job["title"] == record["title"]
    assert job["company"] == record["company"]
    assert job["url"] == record["url"]
    assert job["description_url"] == record["description_url"]
    assert job["platform"] == record["platform"]
    # The date the BD wrote on their sheet, not the day the cycle was built.
    assert job["applied_on"] == record["applied_on"]

    # And it is still all there on the list the waiting-on-a-time table reads.
    waiting = client.get(f"/api/interviews?profile_id={profile}",
                         headers=bd).json()["awaiting_time"]
    row = next(r for r in waiting if r["id"] == made.json()["id"])
    assert row["job"]["applied_on"] == record["applied_on"]
    assert row["job"]["platform"] == record["platform"]


def test_the_applied_date_shown_is_this_profiles_own(client, bd, applied):
    """Two identities can have applied to one posting on different days, and
    showing one of them under the other is worse than showing nothing."""
    mine, theirs = applied["profile"]["id"], applied["other"]["id"]
    shared = [r for r in
              client.get(f"/api/jobs?profile_id={mine}", headers=bd).json()["rows"]]
    others = {r["job_id"]: r for r in
              client.get(f"/api/jobs?profile_id={theirs}", headers=bd).json()["rows"]}
    overlap = next((r for r in shared if r["job_id"] in others), None)
    assert overlap, "the fixture is meant to share postings between the two profiles"

    for profile_id, expected in ((mine, overlap), (theirs, others[overlap["job_id"]])):
        row = client.post("/api/interviews",
                          json={"profile_id": profile_id, "scheduled_at": "",
                                "job_id": overlap["job_id"]}, headers=bd).json()
        assert row["job"]["applied_on"] == expected["applied_on"]


def test_a_take_home_carries_the_whole_posting_too(client, bd, applied):
    """A take-home and the call that produced it describe one conversation.
    Two screens disagreeing about which board it came off, or when the
    application went out, is worse than neither of them saying."""
    profile = applied["profile"]["id"]
    record = next(r for r in
                  client.get(f"/api/jobs?profile_id={profile}", headers=bd).json()["rows"]
                  if r["url"] and r["description_url"])

    made = client.post("/api/assessments",
                       json={"profile_id": profile, "job_id": record["job_id"],
                             "title": "Small RAG pipeline"}, headers=bd)
    assert made.status_code == 201, made.text
    job = made.json()["job"]

    for field in ("title", "company", "url", "description_url", "platform", "applied_on"):
        assert job[field] == record[field], field

    # And still whole on the list every screen reads.
    rows = client.get(f"/api/assessments?profile_id={profile}", headers=bd).json()["rows"]
    row = next(r for r in rows if r["id"] == made.json()["id"])
    assert row["job"]["platform"] == record["platform"]
    assert row["job"]["applied_on"] == record["applied_on"]


def test_a_take_home_set_from_a_call_inherits_that_calls_posting(client, bd, applied):
    """Naming the interview is enough. The call already knows which posting it
    came out of, and asking for it twice is the retyping this all exists to
    stop."""
    profile = applied["profile"]["id"]
    record = next(r for r in
                  client.get(f"/api/jobs?profile_id={profile}", headers=bd).json()["rows"]
                  if r["url"])
    sitting = sat(client, bd, profile, stage="technical", job_id=record["job_id"])

    made = client.post("/api/assessments",
                       json={"profile_id": profile, "interview_id": sitting["id"],
                             "title": "Take-home"}, headers=bd)
    assert made.status_code == 201, made.text
    assert made.json()["job"]["url"] == record["url"]
    assert made.json()["job"]["applied_on"] == record["applied_on"]


def test_the_take_home_applied_date_is_this_profiles_own(client, bd, applied):
    """Same rule as the diary: two identities can have applied on different
    days, and showing one under the other invents a fact."""
    mine, theirs = applied["profile"]["id"], applied["other"]["id"]
    ours = client.get(f"/api/jobs?profile_id={mine}", headers=bd).json()["rows"]
    others = {r["job_id"]: r for r in
              client.get(f"/api/jobs?profile_id={theirs}", headers=bd).json()["rows"]}
    overlap = next((r for r in ours if r["job_id"] in others), None)
    assert overlap, "the fixture is meant to share postings between the two profiles"

    for profile_id, expected in ((mine, overlap), (theirs, others[overlap["job_id"]])):
        row = client.post("/api/assessments",
                          json={"profile_id": profile_id, "job_id": overlap["job_id"],
                                "title": "Shared posting"}, headers=bd).json()
        assert row["job"]["applied_on"] == expected["applied_on"]
